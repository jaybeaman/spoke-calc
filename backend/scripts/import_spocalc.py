#!/usr/bin/env python3
"""
Import rim and hub data from Spocalc Excel spreadsheet.
Download spocalc-2022a.xlsm from https://www.sheldonbrown.com/rinard/spocalc.htm
and place it in this directory before running.
"""

import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

from app.database import SessionLocal, engine, Base
from app.models.rim import Rim
from app.models.hub import Hub

# Create tables if they don't exist
Base.metadata.create_all(bind=engine)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def import_rims_from_excel(db, workbook):
    """Import rims from the Spocalc Excel file."""
    print("Importing rims...")

    # Try to find the rims sheet
    rim_sheet_names = ["Rims", "RIM", "rim", "Rim Database", "RimDB"]
    rim_sheet = None

    for name in rim_sheet_names:
        if name in workbook.sheetnames:
            rim_sheet = workbook[name]
            break

    if not rim_sheet:
        print(f"  Could not find rim sheet. Available sheets: {workbook.sheetnames}")
        return 0

    print(f"  Found rim sheet: {rim_sheet.title}")

    total_imported = 0

    # Find header row and column indices
    header_row = None
    for row_idx, row in enumerate(rim_sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_lower = [str(cell).lower() if cell else "" for cell in row]
        if any("manufacturer" in cell or "brand" in cell or "make" in cell for cell in row_lower):
            header_row = row_idx
            headers = {str(cell).lower() if cell else "": idx for idx, cell in enumerate(row)}
            break

    if not header_row:
        print("  Could not find header row")
        return 0

    # Map column names
    col_map = {}
    for header, idx in headers.items():
        if "manufacturer" in header or "brand" in header or "make" in header:
            col_map["manufacturer"] = idx
        elif "model" in header or "name" in header:
            col_map["model"] = idx
        elif "erd" in header:
            col_map["erd"] = idx
        elif "iso" in header or "bsd" in header or "size" in header:
            col_map["iso_size"] = idx
        elif "inner" in header and "width" in header:
            col_map["inner_width"] = idx
        elif "outer" in header and "width" in header:
            col_map["outer_width"] = idx
        elif "weight" in header:
            col_map["weight"] = idx
        elif "offset" in header:
            col_map["offset"] = idx

    print(f"  Column mapping: {col_map}")

    # Import data rows
    for row in rim_sheet.iter_rows(min_row=header_row + 1, values_only=True):
        try:
            manufacturer = row[col_map.get("manufacturer", 0)] if col_map.get("manufacturer") is not None else None
            model = row[col_map.get("model", 1)] if col_map.get("model") is not None else None
            erd = row[col_map.get("erd", 2)] if col_map.get("erd") is not None else None

            if not manufacturer or not model or not erd:
                continue

            # Clean up values
            manufacturer = str(manufacturer).strip()
            model = str(model).strip()

            try:
                erd = float(erd)
            except (ValueError, TypeError):
                continue

            if erd < 200 or erd > 700:  # Sanity check
                continue

            # Check if already exists
            existing = db.query(Rim).filter(
                Rim.manufacturer == manufacturer,
                Rim.model == model,
                Rim.erd == erd
            ).first()

            if existing:
                continue

            # Parse optional fields
            iso_size = None
            if col_map.get("iso_size") is not None:
                try:
                    iso_size = int(float(row[col_map["iso_size"]]))
                except (ValueError, TypeError):
                    pass

            inner_width = None
            if col_map.get("inner_width") is not None:
                try:
                    inner_width = float(row[col_map["inner_width"]])
                except (ValueError, TypeError):
                    pass

            outer_width = None
            if col_map.get("outer_width") is not None:
                try:
                    outer_width = float(row[col_map["outer_width"]])
                except (ValueError, TypeError):
                    pass

            weight = None
            if col_map.get("weight") is not None:
                try:
                    weight = float(row[col_map["weight"]])
                except (ValueError, TypeError):
                    pass

            offset = 0
            if col_map.get("offset") is not None:
                try:
                    offset = float(row[col_map["offset"]])
                except (ValueError, TypeError):
                    pass

            rim = Rim(
                manufacturer=manufacturer,
                model=model,
                erd=erd,
                iso_size=iso_size,
                inner_width=inner_width,
                outer_width=outer_width,
                weight=weight,
                drilling_offset=offset,
                is_reference=True
            )
            db.add(rim)
            total_imported += 1

        except Exception as e:
            continue

    db.commit()
    print(f"  Imported {total_imported} rims")
    return total_imported


def import_hubs_from_excel(db, workbook):
    """Import hubs from the Spocalc Excel file."""
    print("Importing hubs...")

    # Try to find the hubs sheet
    hub_sheet_names = ["Hubs", "HUB", "hub", "Hub Database", "HubDB"]
    hub_sheet = None

    for name in hub_sheet_names:
        if name in workbook.sheetnames:
            hub_sheet = workbook[name]
            break

    if not hub_sheet:
        print(f"  Could not find hub sheet. Available sheets: {workbook.sheetnames}")
        return 0

    print(f"  Found hub sheet: {hub_sheet.title}")

    total_imported = 0

    # Find header row
    header_row = None
    for row_idx, row in enumerate(hub_sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_lower = [str(cell).lower() if cell else "" for cell in row]
        if any("manufacturer" in cell or "brand" in cell or "make" in cell for cell in row_lower):
            header_row = row_idx
            headers = {str(cell).lower() if cell else "": idx for idx, cell in enumerate(row)}
            break

    if not header_row:
        print("  Could not find header row")
        return 0

    # Map column names - hub measurements are trickier
    col_map = {}
    for header, idx in headers.items():
        if "manufacturer" in header or "brand" in header or "make" in header:
            col_map["manufacturer"] = idx
        elif "model" in header or "name" in header:
            col_map["model"] = idx
        elif "flange" in header and ("dia" in header or "pcd" in header):
            if "left" in header or "l " in header:
                col_map["flange_dia_left"] = idx
            elif "right" in header or "r " in header:
                col_map["flange_dia_right"] = idx
            else:
                col_map["flange_dia"] = idx
        elif "offset" in header or "center" in header or "dish" in header:
            if "left" in header or "l " in header:
                col_map["offset_left"] = idx
            elif "right" in header or "r " in header:
                col_map["offset_right"] = idx
        elif "spoke" in header and "hole" in header:
            col_map["spoke_holes"] = idx
        elif "position" in header or "front" in header or "rear" in header:
            col_map["position"] = idx

    print(f"  Column mapping: {col_map}")

    # Import data rows
    for row in hub_sheet.iter_rows(min_row=header_row + 1, values_only=True):
        try:
            manufacturer = row[col_map.get("manufacturer", 0)] if col_map.get("manufacturer") is not None else None
            model = row[col_map.get("model", 1)] if col_map.get("model") is not None else None

            if not manufacturer or not model:
                continue

            manufacturer = str(manufacturer).strip()
            model = str(model).strip()

            # Check if already exists
            existing = db.query(Hub).filter(
                Hub.manufacturer == manufacturer,
                Hub.model == model
            ).first()

            if existing:
                continue

            # Parse flange diameter
            flange_dia_left = None
            flange_dia_right = None

            if col_map.get("flange_dia") is not None:
                try:
                    flange_dia = float(row[col_map["flange_dia"]])
                    flange_dia_left = flange_dia
                    flange_dia_right = flange_dia
                except (ValueError, TypeError):
                    pass

            if col_map.get("flange_dia_left") is not None:
                try:
                    flange_dia_left = float(row[col_map["flange_dia_left"]])
                except (ValueError, TypeError):
                    pass

            if col_map.get("flange_dia_right") is not None:
                try:
                    flange_dia_right = float(row[col_map["flange_dia_right"]])
                except (ValueError, TypeError):
                    pass

            # Parse offsets
            offset_left = None
            offset_right = None

            if col_map.get("offset_left") is not None:
                try:
                    offset_left = float(row[col_map["offset_left"]])
                except (ValueError, TypeError):
                    pass

            if col_map.get("offset_right") is not None:
                try:
                    offset_right = float(row[col_map["offset_right"]])
                except (ValueError, TypeError):
                    pass

            # Skip if we don't have enough data
            if flange_dia_left is None and offset_left is None:
                continue

            # Parse spoke holes
            spoke_count = None
            if col_map.get("spoke_holes") is not None:
                try:
                    spoke_count = int(float(row[col_map["spoke_holes"]]))
                except (ValueError, TypeError):
                    pass

            # Parse position
            position = None
            if col_map.get("position") is not None:
                pos_val = str(row[col_map["position"]]).lower()
                if "front" in pos_val:
                    position = "front"
                elif "rear" in pos_val:
                    position = "rear"

            hub = Hub(
                manufacturer=manufacturer,
                model=model,
                position=position,
                flange_diameter_left=flange_dia_left or 0,
                flange_diameter_right=flange_dia_right or flange_dia_left or 0,
                flange_offset_left=offset_left or 0,
                flange_offset_right=offset_right or offset_left or 0,
                spoke_count=spoke_count,
                is_reference=True
            )
            db.add(hub)
            total_imported += 1

        except Exception as e:
            continue

    db.commit()
    print(f"  Imported {total_imported} hubs")
    return total_imported


def main():
    # Look for the Excel file
    excel_files = [
        os.path.join(SCRIPT_DIR, "spocalc-2022a.xlsm"),
        os.path.join(SCRIPT_DIR, "spocalc.xls"),
        os.path.join(SCRIPT_DIR, "spocalc.xlsm"),
    ]

    excel_file = None
    for f in excel_files:
        if os.path.exists(f):
            excel_file = f
            break

    if not excel_file:
        print("Error: Could not find Spocalc Excel file.")
        print("Please download from: https://www.sheldonbrown.com/rinard/spocalc.htm")
        print(f"And place it in: {SCRIPT_DIR}")
        return

    print(f"Loading {excel_file}...")

    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    print(f"Available sheets: {workbook.sheetnames}")

    db = SessionLocal()

    try:
        rims_count = import_rims_from_excel(db, workbook)
        hubs_count = import_hubs_from_excel(db, workbook)

        print(f"\nImport complete!")
        print(f"  Total rims in database: {db.query(Rim).count()}")
        print(f"  Total hubs in database: {db.query(Hub).count()}")

    finally:
        db.close()


if __name__ == "__main__":
    main()
